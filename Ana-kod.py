# excel dosyasındaki verileri gün ve saate göre seçip yazdıran program

import datetime
from tkinter import *
from time import strftime
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook,load_workbook
from tkinter import ttk
from time import *
from tkcalendar import DateEntry


wb = None
ws = None

master = Tk()
master.title("Kullanıcı Arayüzü") #Master penceresine isim verme
canvas = Canvas(master, height=800, width=1800) # Ana Pencere Boyutu ayarlama
canvas.pack()           # pack place grid 3 tane ekleme seçeneği var


#3  tane frame yapısından oluşan bir arayüz tasarladık

frame_ust = Frame(master, bg='#add8e6')
frame_ust.place(relx=0.1, rely=0.12, relwidth=0.8, relheight=0.1) # rel x ve rel y de framenin_üst nerden başlıcağını gösteriyor
frame_alt_sol = Frame(master, bg='#add8e6')
frame_alt_sol.place(relx=0.1, rely=0.25, relwidth=0.2, relheight=0.7)
frame_alt_sag = Frame(master, bg='#add8e6')
frame_alt_sag.place(relx=0.32, rely=0.25, relwidth=0.58, relheight=0.7)


def aktif_saat():
    # aktif bir saat fonksiyonu tanımlayarak saat verisini ekrana yazdırmak için kullanıyoruz1
    time_string = strftime("%d/%m/%Y\n%H:%M:%S")
    tarih_saat.config(text=time_string)
    tarih_saat.after(1000,aktif_saat)

# Tarih ve saat
tarih_saat = Label(frame_ust,bg='#add8e6', font="verdana 9 bold")
tarih_saat.pack(pady=10,padx=10, side=RIGHT)
tarih = Label(frame_ust,bg='#add8e6',text="Tarih:\nSaat:", font="verdana 9 bold")
tarih.pack(pady=10,padx=10, side=RIGHT)
kalan_gün_label = Label(frame_ust,bg='#add8e6',text="Kalan Gün:", font="verdana 9 bold")
kalan_gün_label.pack(pady=5,padx=5, side=LEFT)
kalan_gün= Label(frame_ust, bg='#add8e6',font="verdana 9 bold")
kalan_gün.pack(pady=5, padx=5, side=LEFT )

def dosya_sec():
    global wb, ws

    wb = load_workbook(tk.filedialog.askopenfilename()) #excel dosyasını seçme
    ws = wb.active

    scrollbarx= Scrollbar(frame_alt_sag,orient=HORIZONTAL)    # yatay ve dikey kaydırma cubuğunu tanımlama
    scrollbary = Scrollbar(frame_alt_sag, orient=VERTICAL)

    # Excel 1 satırın 1 hücresinden 25 hücesine kadar verileri tablonun yatay ekseni kabul ediliyor.
    l1 = ws.iter_rows(min_row=1, max_row=1, max_col=25,values_only=True)
    # aynı şekilde 1 sütünün 2. hücresinden 250. Hücresine kadarı sütün kabul ediyoruz.
    r_set = ws.iter_rows(min_row=2, max_row=240, max_col=25, values_only=True)
    l1 = [r for r in l1]
    r_set = [r for r in r_set]

    trv = ttk.Treeview(frame_alt_sag,selectmode='extended')  # Oluşacak tablonunn nerde listeleneceği bilgileri
    trv.place(relx=0.01,rely=0.1,width=825,height=465)

    style = ttk.Style()
    style.configure('trv', background='blue')

    trv.configure(yscrollcommand=scrollbary.set,xscrollcommand=scrollbarx.set,)# tablo ile kaydırma cubuklarını eşleştirme
    scrollbary.configure(command=trv.yview)
    scrollbarx.configure(command=trv.xview)

    scrollbary.place(relx=0.934,rely=0.1, width=22,height=460)          # kaydırma cubuklarının konumunu belirleme
    scrollbarx.place(relx=0.01, rely=0.922, width=800, height=22)


    trv['show'] = 'headings'
    trv['columns'] = l1[0]

    for i in  l1[0]:     # excel tablosundaki verilerin arayüzdeki tabloya çekilmesi
        trv.column(i,width=100,anchor='c')

    for i in  l1[0]:
        trv.heading(i,text=i)

    for dt in r_set:
        trv.insert("", 'end', iid=dt[0], values=dt )

def dosya_verileri_pencere():

    pencere2 = tk.Tk()
    pencere2.geometry("380x200")
    Bitiş_tarihi = Label(pencere2, text="Bitiş tarihi Seç:" ,font="verdana 9 bold")
    Bitiş_tarihi.place(relx=0.01, rely=0.1, relwidth=0.3, relheight=0.1)

    cal = DateEntry(pencere2, selectmode='day') # takvim oluşturma

    cal.place(relx=0.3, rely=0.099, relwidth=0.2, relheight=0.1)




    def tarih_onayla():

        secilen_tarih = cal.get_date()
        with open("secilen_tarih.txt", "w") as file:  # seçilen tarihi txt dosyasına kaydet
            file.write(str(secilen_tarih))
        secilen_yıl = secilen_tarih.year
        secilen_ay = secilen_tarih.month
        secilen_gun = secilen_tarih.day


        # Tarih farkını alma
        delta = datetime.datetime(secilen_yıl, secilen_ay, secilen_gun) - datetime.datetime.now()

        days = delta.days
        # Aktif olarak kaç gün kaldığını yenileme
        hours, rem = divmod(delta.seconds, 3600)
        minutes, seconds = divmod(rem, 60)
        kalan_gün.after(1000, tarih_onayla)

        kalan_gün_label3= Label(pencere2,text="Proje Süresi: ",font="verdana 9 bold")
        kalan_gün_label3.place(relx=0.01, rely=0.2 ,relwidth=0.27, relheight=0.18)
        kalan_gün_cevap3= Label(pencere2,text=[days,"GÜN", hours,"SAAT",minutes,"DAKİKA"],font="verdana 9 bold")
        kalan_gün_cevap3.place(relx=0.25, rely=0.24 ,relwidth=0.8, relheight=0.1)

    l1 = tk.Label(pencere2, text='Onaylanan Tarih')  # Label to display date
    l1.place(relx=0.72, rely=0.099, relwidth=0.3, relheight=0.1)
    b1 = tk.Button(pencere2, text='Tarihi Onayla', command=tarih_onayla)
    b1.place(relx=0.53, rely=0.099, relwidth=0.2, relheight=0.1)
    b2 = tk.Button(pencere2, text='Kaydet',command=pencere2.destroy)
    b2.place(relx=0.4, rely=0.7, relwidth=0.2, relheight=0.1)

    pencere2.mainloop()

saat_sayici=0
gun_sayici=1

def sayac():
    global saat_sayici, gun_sayici
    saat_sayici += 1

    if saat_sayici == 24:
        saat_sayici = 0
        gun_sayici += 1
    print("s", saat_sayici, "g", gun_sayici)
    master.after(3600*1000,sayac)



def veri_cekme():
    global saat_sayici, gun_sayici, wb, ws



    sure = "{:d}.GÜN".format(gun_sayici)  # kaçıncı gün olduğunu excel sütün indeksine cevirme
    now = datetime.datetime.now()
    hour = now.hour
    rounded_hour = round(hour)  # saati tam sayıya yuvarlama

    saat = "{:02d}.00".format(rounded_hour)  # saat formatını excel satır indeksine cevirme
    columns = ws.columns
    column_names = [column[0].value for column in columns]
    row_index = column_names.index(saat)

    rows = ws.rows
    row_values = [row[0].value for row in rows]
    hour_index = row_values.index(sure)
    # kesişim hücresini çekme işlemi
    cell = ws.cell(row=hour_index + 1, column=row_index + 1).value
    cell1 = ws.cell(row=hour_index + 2, column=row_index + 1).value
    print("AKTİF GRUP:", cell, " /", " Işık Şiddeti: ", cell1)

    with open("secilen_tarih.txt", "r") as file:
        tarih_str = file.read()
        tarih = datetime.datetime.strptime(tarih_str,"%Y-%m-%d")
        bitiş_tarihi = tarih.date()

        today = datetime.datetime.today()
        bugun_tarihi= today.strftime("%Y-%m-%d")

        bitiş_tarihi_str = bitiş_tarihi.strftime("%Y-%m-%d")

        if bitiş_tarihi_str==bugun_tarihi:
            pencere3 = Tk()
            pencere3.title("Kullanıcı Arayüzü")
            pencere3.title("Bilgi Penceresi")
            etiket = tk.Label(pencere3, text="Program tamamlandı.")
            etiket.pack()

            b3 = tk.Button(pencere3, text='Kaydet', command=pencere3.destroy)
            b3.place(relx=0.4, rely=0.7, relwidth=0.2, relheight=0.1)
            pencere3.mainloop()

        else:
            print(saat_sayici,gun_sayici)
            # While dongüsü çalışırken arayüz genel kod çalışmaya ara veriyor ve büyüzden aratyüz yanıt alınamadı hatası veriyor
            master.after(60 * 1000,veri_cekme)
            # bu yüzden döngü yerine after seçeneği kullanıldı.

def program_başlat():
        veri_cekme()





Dosya_secme_butonu=Button(frame_alt_sol, text="Dosya seç",height= 5, width=15,command=dosya_sec).pack(pady=10,padx=10, )
Dosya_bilgileri_ayarla=Button(frame_alt_sol, text="Program Veriler Giriş",height= 5, width=15,command=dosya_verileri_pencere).pack(pady=10,padx=10, )
Başlatma=Button(frame_alt_sol, text="programı başlat",height= 5, width=15,command=lambda:[program_başlat(),sayac()]).pack(pady=10,padx=10 )
Manuel_Mod=Button(frame_alt_sol, text="Manuel Mod",height= 5, width=15,).pack(pady=10,padx=10 ) # daha atanmadı



aktif_saat()
master.mainloop()



