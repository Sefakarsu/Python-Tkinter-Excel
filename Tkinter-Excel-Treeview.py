
# Excel Dosyasınıdaki verileri treeview ile arayüze yazdırıp scrollbar ekleme
import datetime
from tkinter import *
from time import strftime
import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook,load_workbook
from tkinter import ttk
from time import *


wb = None
ws = None

master = Tk()
master.title("Kullanıcı Arayüzü") #Master penceresine isim verme
canvas = Canvas(master, height=800, width=1800) # Ana Pencere Boyutu ayarlama
canvas.pack()           # pack place grid 3 tane ekleme seçeneği var


#3  tane frame yapısından oluşan bir arayüz tasarladık

frame_ust = Frame(master, bg='#add8e6')
frame_ust.place(relx=0.1, rely=0.12, relwidth=0.8, relheight=0.1) # rel x ve rel y de framenin_üst nerden başlıcağını gösteriyor
frame_alt_sol = Frame(master, bg='#add8e6')
frame_alt_sol.place(relx=0.1, rely=0.25, relwidth=0.2, relheight=0.7)
frame_alt_sag = Frame(master, bg='#add8e6')
frame_alt_sag.place(relx=0.32, rely=0.25, relwidth=0.58, relheight=0.7)


def aktif_saat():
    # aktif bir saat fonksiyonu tanımlayarak saat verisini ekrana yazdırmak için kullanıyoruz1
    time_string = strftime("%d/%m/%Y\n%H:%M:%S")
    tarih_saat.config(text=time_string)
    tarih_saat.after(1000,aktif_saat)

# Tarih ve saat
tarih_saat = Label(frame_ust,bg='#add8e6', font="verdana 9 bold")
tarih_saat.pack(pady=10,padx=10, side=RIGHT)
tarih = Label(frame_ust,bg='#add8e6',text="Tarih:\nSaat:", font="verdana 9 bold")
tarih.pack(pady=10,padx=10, side=RIGHT)


def dosya_sec():
    global wb, ws

    wb = load_workbook(tk.filedialog.askopenfilename()) #excel dosyasını seçme
    ws = wb.active

    scrollbarx= Scrollbar(frame_alt_sag,orient=HORIZONTAL)    # yatay ve dikey kaydırma cubuğunu tanımlama
    scrollbary = Scrollbar(frame_alt_sag, orient=VERTICAL)

    # Excel 1 satırın 1 hücresinden 25 hücesine kadar verileri tablonun yatay ekseni kabul ediliyor.
    l1 = ws.iter_rows(min_row=1, max_row=1, max_col=25,values_only=True)
    # aynı şekilde 1 sütünün 2. hücresinden 250. Hücresine kadarı sütün kabul ediyoruz.
    r_set = ws.iter_rows(min_row=2, max_row=240, max_col=25, values_only=True)
    l1 = [r for r in l1]
    r_set = [r for r in r_set]

    trv = ttk.Treeview(frame_alt_sag,selectmode='extended')  # Oluşacak tablonunn nerde listeleneceği bilgileri
    trv.place(relx=0.01,rely=0.1,width=825,height=465)

    style = ttk.Style()
    style.configure('trv', background='blue')

    trv.configure(yscrollcommand=scrollbary.set,xscrollcommand=scrollbarx.set,)# tablo ile kaydırma cubuklarını eşleştirme
    scrollbary.configure(command=trv.yview)
    scrollbarx.configure(command=trv.xview)

    scrollbary.place(relx=0.934,rely=0.1, width=22,height=460)          # kaydırma cubuklarının konumunu belirleme
    scrollbarx.place(relx=0.01, rely=0.922, width=800, height=22)


    trv['show'] = 'headings'
    trv['columns'] = l1[0]

    for i in  l1[0]:     # excel tablosundaki verilerin arayüzdeki tabloya çekilmesi
        trv.column(i,width=100,anchor='c')

    for i in  l1[0]:
        trv.heading(i,text=i)

    for dt in r_set:
        trv.insert("", 'end', iid=dt[0], values=dt )

Dosya_secme_butonu=Button(frame_alt_sol, text="Dosya seç",height= 5, width=15,command=dosya_sec).pack(pady=10,padx=10, )
Dosya_bilgileri_ayarla=Button(frame_alt_sol, text="Program Veriler Giriş",height= 5, width=15).pack(pady=10,padx=10, )
Başlatma=Button(frame_alt_sol, text="programı başlat",height= 5, width=15).pack(pady=10,padx=10 )
Manuel_Mod=Button(frame_alt_sol, text="Manuel Mod",height= 5, width=15).pack(pady=10,padx=10 ) # daha atanmadı

aktif_saat()
master.mainloop()